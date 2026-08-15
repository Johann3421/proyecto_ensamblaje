import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import List, Dict

def generate_checklist_excel(model_name: str, items: List[Dict]) -> bytes:
    """Genera un archivo Excel (.xlsx) estilizado con el formato oficial de QC KENYA"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"QC_{model_name[:20]}"
    
    # Encabezado principal
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"CHECKLIST DE CONTROL DE CALIDAD – PC KENYA {model_name.upper()}"
    title_cell.font = Font(name="Segoe UI", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="0078D4", end_color="0078D4", fill_type="solid") # Microsoft Blue
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 35

    # Headers de columnas
    headers = [
        ("A2", "Paso_Nro", 10),
        ("B2", "Operacion", 35),
        ("C2", "Descripcion_Detallada", 45),
        ("D2", "Criterio_Control_Calidad", 45),
        ("E2", "Multimedia_URL_O_Nombre", 30),
    ]

    header_fill = PatternFill(start_color="F3F2F1", end_color="F3F2F1", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="323130")
    thin_border = Border(
        left=Side(style="thin", color="E1DFDD"),
        right=Side(style="thin", color="E1DFDD"),
        top=Side(style="thin", color="E1DFDD"),
        bottom=Side(style="thin", color="E1DFDD")
    )

    for cell_ref, text, col_width in headers:
        cell = ws[cell_ref]
        cell.value = text
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center" if "Nro" in text else "left", vertical="center")
        cell.border = thin_border
        col_letter = cell_ref[0]
        ws.column_dimensions[col_letter].width = col_width

    ws.row_dimensions[2].height = 25

    # Llenar datos
    row_num = 3
    for it in items:
        ws[f"A{row_num}"] = it.get("step_number", row_num - 2)
        ws[f"B{row_num}"] = it.get("operation", "")
        ws[f"C{row_num}"] = it.get("description", "")
        ws[f"D{row_num}"] = it.get("qc_criteria", "")
        ws[f"E{row_num}"] = it.get("media_url", "")

        for col in ["A", "B", "C", "D", "E"]:
            c = ws[f"{col}{row_num}"]
            c.font = Font(name="Segoe UI", size=10)
            c.border = thin_border
            c.alignment = Alignment(
                horizontal="center" if col == "A" else "left",
                vertical="center",
                wrap_text=True
            )
        ws.row_dimensions[row_num].height = 30
        row_num += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def parse_checklist_excel(file_bytes: bytes) -> List[Dict]:
    """Parsea un archivo Excel subido y extrae la lista de pasos"""
    wb = openpyxl.load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    
    items = []
    # Buscar la fila de encabezados
    start_row = 3
    # Si la fila 2 tiene texto de cabecera, los datos empiezan en fila 3
    for r in range(start_row, ws.max_row + 1):
        step_num_val = ws.cell(row=r, column=1).value
        operation_val = ws.cell(row=r, column=2).value
        desc_val = ws.cell(row=r, column=3).value
        criteria_val = ws.cell(row=r, column=4).value
        media_val = ws.cell(row=r, column=5).value

        if not operation_val and not step_num_val:
            continue

        try:
            step_num = int(step_num_val) if step_num_val is not None else len(items) + 1
        except (ValueError, TypeError):
            step_num = len(items) + 1

        items.append({
            "step_number": step_num,
            "operation": str(operation_val or "").strip(),
            "description": str(desc_val or "").strip(),
            "qc_criteria": str(criteria_val or "").strip(),
            "media_url": str(media_val or "").strip(),
            "media_type": "gif" if "gif" in str(media_val or "").lower() else "image"
        })
    return items
